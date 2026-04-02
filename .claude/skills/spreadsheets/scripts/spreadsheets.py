#!/usr/bin/env python3
"""Spreadsheet helper for Claude Code: read, write, edit, and convert CSV/XLSX files."""

import argparse
import csv
import json
import sys
from pathlib import Path


def is_xlsx(path: str) -> bool:
    return Path(path).suffix.lower() in (".xlsx", ".xls")


def parse_col_ref(ref: str) -> int:
    """Convert a column reference to 0-based index. Accepts 'A'-'ZZ' or numeric."""
    if ref.isdigit():
        return int(ref)
    ref = ref.upper()
    result = 0
    for ch in ref:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def parse_cell_ref(ref: str):
    """Parse A1-notation cell ref into (row_0based, col_0based)."""
    i = 0
    while i < len(ref) and ref[i].isalpha():
        i += 1
    col_str = ref[:i]
    row_str = ref[i:]
    if not col_str or not row_str:
        raise ValueError(f"Invalid cell reference: {ref}")
    return int(row_str) - 1, parse_col_ref(col_str)


def parse_range(range_str: str):
    """Parse 'start:end' into (start, end) inclusive, 0-based. Input is 1-based."""
    parts = range_str.split(":")
    if len(parts) != 2:
        raise ValueError(f"Invalid range: {range_str}. Expected start:end")
    return int(parts[0]) - 1, int(parts[1]) - 1


def parse_col_range(range_str: str):
    """Parse column range like 'A:C' or '0:2' into (start, end) inclusive, 0-based."""
    parts = range_str.split(":")
    if len(parts) != 2:
        raise ValueError(f"Invalid column range: {range_str}")
    return parse_col_ref(parts[0]), parse_col_ref(parts[1])


# -- Read ------------------------------------------------------------------

def read_csv(path, row_range=None, col_range=None):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    return _slice_rows(rows, row_range, col_range)


def read_xlsx(path, sheet=None, row_range=None, col_range=None):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        ws = wb.active
    elif sheet.isdigit():
        ws = wb.worksheets[int(sheet)]
    else:
        ws = wb[sheet]

    rows = []
    for row in ws.iter_rows():
        rows.append([_cell_value(c) for c in row])
    wb.close()
    return _slice_rows(rows, row_range, col_range)


def _cell_value(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v)


def _slice_rows(rows, row_range, col_range):
    if row_range:
        start, end = row_range
        rows = rows[start : end + 1]
    if col_range:
        cs, ce = col_range
        rows = [row[cs : ce + 1] for row in rows]
    return rows


def format_table(rows):
    if not rows:
        return "(empty)"
    # Compute column widths
    ncols = max(len(r) for r in rows)
    widths = [0] * ncols
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))
    # Cap widths at 40 to keep output readable
    widths = [min(w, 40) for w in widths]
    lines = []
    for ri, row in enumerate(rows):
        parts = []
        for i in range(ncols):
            val = row[i] if i < len(row) else ""
            if len(val) > 40:
                val = val[:37] + "..."
            parts.append(val.ljust(widths[i]))
        lines.append(" | ".join(parts))
        if ri == 0:
            lines.append("-+-".join("-" * w for w in widths))
    return "\n".join(lines)


def cmd_read(args):
    row_range = parse_range(args.rows) if args.rows else None
    col_range = parse_col_range(args.cols) if args.cols else None

    if is_xlsx(args.file):
        rows = read_xlsx(args.file, args.sheet, row_range, col_range)
    else:
        rows = read_csv(args.file, row_range, col_range)

    print(format_table(rows))


# -- Write -----------------------------------------------------------------

def cmd_write(args):
    data = json.load(sys.stdin)
    headers = data.get("headers", [])
    rows = data.get("rows", [])

    if is_xlsx(args.output):
        write_xlsx(args.output, headers, rows, args.sheet or "Sheet1")
    else:
        write_csv(args.output, headers, rows)

    print(f"Wrote {len(rows)} rows to {args.output}", file=sys.stderr)


def write_csv(path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if headers:
            writer.writerow(headers)
        writer.writerows(rows)


def write_xlsx(path, headers, rows, sheet_name):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


# -- Edit ------------------------------------------------------------------

def cmd_edit(args):
    if not args.set:
        print("Error: --set is required for edit", file=sys.stderr)
        sys.exit(1)

    edits = {}
    for s in args.set:
        if "=" not in s:
            print(f"Error: invalid --set format: {s}. Expected CELL=VALUE", file=sys.stderr)
            sys.exit(1)
        cell, value = s.split("=", 1)
        row, col = parse_cell_ref(cell)
        edits[(row, col)] = value

    if is_xlsx(args.file):
        edit_xlsx(args.file, args.sheet, edits)
    else:
        edit_csv(args.file, edits)

    print(f"Applied {len(edits)} edit(s) to {args.file}", file=sys.stderr)


def edit_csv(path, edits):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    # Expand grid if needed
    max_row = max(r for r, c in edits)
    max_col = max(c for r, c in edits)
    while len(rows) <= max_row:
        rows.append([])
    for row in rows:
        while len(row) <= max_col:
            row.append("")

    for (r, c), val in edits.items():
        rows[r][c] = val

    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def edit_xlsx(path, sheet, edits):
    import openpyxl

    p = Path(path)
    if p.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()

    if sheet is None:
        ws = wb.active
    elif sheet.isdigit():
        ws = wb.worksheets[int(sheet)]
    else:
        ws = wb[sheet]

    for (r, c), val in edits.items():
        ws.cell(row=r + 1, column=c + 1, value=val)

    wb.save(path)


# -- Convert ---------------------------------------------------------------

def cmd_convert(args):
    if is_xlsx(args.input):
        rows = read_xlsx(args.input, args.sheet)
    else:
        rows = read_csv(args.input)

    if is_xlsx(args.output):
        # First row as headers if present
        headers = rows[0] if rows else []
        data_rows = rows[1:] if rows else []
        write_xlsx(args.output, headers, data_rows, args.sheet or "Sheet1")
    else:
        write_csv_rows(args.output, rows)

    print(f"Converted {args.input} -> {args.output} ({len(rows)} rows)", file=sys.stderr)


def write_csv_rows(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


# -- CLI -------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Spreadsheet helper")
    sub = parser.add_subparsers(dest="command", required=True)

    # read
    p_read = sub.add_parser("read", help="Read a spreadsheet")
    p_read.add_argument("file")
    p_read.add_argument("--sheet", default=None)
    p_read.add_argument("--rows", default=None, help="Row range, 1-based inclusive (e.g. 1:10)")
    p_read.add_argument("--cols", default=None, help="Col range (e.g. A:C or 0:2)")
    p_read.set_defaults(func=cmd_read)

    # write
    p_write = sub.add_parser("write", help="Write a new spreadsheet from JSON stdin")
    p_write.add_argument("output")
    p_write.add_argument("--sheet", default=None)
    p_write.set_defaults(func=cmd_write)

    # edit
    p_edit = sub.add_parser("edit", help="Edit cells in a spreadsheet")
    p_edit.add_argument("file")
    p_edit.add_argument("--sheet", default=None)
    p_edit.add_argument("--set", action="append", help="CELL=VALUE (e.g. B3=hello)")
    p_edit.set_defaults(func=cmd_edit)

    # convert
    p_convert = sub.add_parser("convert", help="Convert between CSV and XLSX")
    p_convert.add_argument("input")
    p_convert.add_argument("output")
    p_convert.add_argument("--sheet", default=None)
    p_convert.set_defaults(func=cmd_convert)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
